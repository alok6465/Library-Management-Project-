from flask import Flask, render_template, request, send_file, jsonify
import csv
import json
import os
from io import BytesIO, StringIO
# Removed reportlab imports since we're using DOC format now

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

def read_excel_simple(filepath):
    """Simple Excel reader using openpyxl"""
    from openpyxl import load_workbook
    wb = load_workbook(filepath)
    ws = wb.active
    data = []
    for row in ws.iter_rows(values_only=True):
        data.append([str(cell) if cell is not None else '' for cell in row])
    return data

def read_csv_simple(filepath):
    """Simple CSV reader"""
    data = []
    with open(filepath, 'r', encoding='utf-8') as file:
        reader = csv.reader(file)
        for row in reader:
            data.append(row)
    return data

def data_to_html_table(data):
    """Convert data to HTML table"""
    if not data:
        return "<p>No data found</p>"
    
    html = '<table class="table table-striped table-hover" id="dataTable">\n'
    
    # Header
    if data:
        html += '<thead><tr>'
        for cell in data[0]:
            html += f'<th>{cell}</th>'
        html += '</tr></thead>\n'
    
    # Body
    html += '<tbody>'
    for row in data[1:]:
        html += '<tr>'
        for cell in row:
            html += f'<td>{cell}</td>'
        html += '</tr>'
    html += '</tbody></table>'
    
    return html

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_file():
    if 'file' not in request.files:
        return jsonify({'error': 'No file selected'})
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'})
    
    if file and (file.filename.endswith('.xlsx') or file.filename.endswith('.csv')):
        filename = file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)
        
        try:
            if filename.endswith('.xlsx'):
                data = read_excel_simple(filepath)
            else:
                data = read_csv_simple(filepath)
            
            table_html = data_to_html_table(data)
            
            return jsonify({
                'success': True,
                'table': table_html,
                'filename': filename,
                'rows': len(data) - 1 if data else 0,
                'columns': len(data[0]) if data else 0
            })
        except Exception as e:
            return jsonify({'error': f'Error reading file: {str(e)}'})
    
    return jsonify({'error': 'Invalid file format. Only .xlsx and .csv files are allowed.'})

@app.route('/download_html', methods=['POST'])
def download_html():
    filename = request.json.get('filename')
    search_term = request.json.get('search', '').lower()
    
    if not filename:
        return jsonify({'error': 'No filename provided'})
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    try:
        if filename.endswith('.xlsx'):
            data = read_excel_simple(filepath)
        else:
            data = read_csv_simple(filepath)
        
        # Apply search filter
        if search_term and data:
            filtered_data = [data[0]]  # Keep header
            for row in data[1:]:
                if any(search_term in str(cell).lower() for cell in row):
                    filtered_data.append(row)
            data = filtered_data
        
        # Create HTML with enhanced styling and titles
        html_content = f'''
<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>📊 Data Export - {filename}</title>
    <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.1.3/dist/css/bootstrap.min.css" rel="stylesheet">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
    <style>
        * {{ box-sizing: border-box; }}
        body {{ 
            font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif; 
            margin: 0; 
            padding: 20px;
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            min-height: 100vh;
        }}
        .container {{ 
            background: white; 
            border-radius: 15px; 
            box-shadow: 0 20px 40px rgba(0,0,0,0.1);
            padding: 30px;
            margin: 0 auto;
            max-width: 95vw;
        }}
        .header {{ 
            text-align: center;
            margin-bottom: 30px;
            padding: 20px;
            background: linear-gradient(135deg, #4facfe 0%, #00f2fe 100%);
            border-radius: 10px;
            color: white;
            box-shadow: 0 10px 30px rgba(79, 172, 254, 0.3);
        }}
        .header h1 {{ 
            font-size: 2.5rem;
            font-weight: 700;
            margin: 0;
            text-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }}
        .header p {{ 
            font-size: 1.1rem;
            margin: 10px 0 0 0;
            opacity: 0.9;
        }}
        .stats-bar {{ 
            display: flex;
            justify-content: space-around;
            margin: 20px 0;
            padding: 15px;
            background: #f8f9fa;
            border-radius: 10px;
            border-left: 4px solid #007bff;
        }}
        .stat-item {{ 
            text-align: center;
        }}
        .stat-number {{ 
            font-size: 1.8rem;
            font-weight: 700;
            color: #007bff;
        }}
        .stat-label {{ 
            font-size: 0.9rem;
            color: #6c757d;
            text-transform: uppercase;
            letter-spacing: 0.5px;
        }}
        .table-container {{ 
            overflow: auto; 
            max-height: 70vh; 
            border: 2px solid #e9ecef;
            border-radius: 10px;
            box-shadow: 0 5px 15px rgba(0,0,0,0.08);
            background: white;
        }}
        table {{ 
            width: 100%; 
            border-collapse: collapse; 
            min-width: max-content;
            font-size: 0.9rem;
        }}
        th {{ 
            background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
            color: white; 
            padding: 15px 12px; 
            text-align: left; 
            position: sticky; 
            top: 0; 
            z-index: 10;
            white-space: nowrap;
            font-weight: 600;
            text-transform: uppercase;
            letter-spacing: 0.5px;
            font-size: 0.8rem;
            border-right: 1px solid rgba(255,255,255,0.2);
        }}
        td {{ 
            padding: 12px; 
            border-bottom: 1px solid #e9ecef;
            white-space: nowrap;
            transition: background-color 0.2s ease;
        }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        tr:hover {{ 
            background-color: #e3f2fd !important;
            transform: scale(1.001);
            box-shadow: 0 2px 8px rgba(0,0,0,0.1);
        }}
        .summary {{ 
            margin-top: 30px; 
            padding: 20px; 
            background: linear-gradient(135deg, #a8edea 0%, #fed6e3 100%);
            border-radius: 10px;
            border-left: 4px solid #28a745;
        }}
        .summary h5 {{ 
            color: #155724;
            font-weight: 600;
            margin-bottom: 15px;
        }}
        .filter-badge {{ 
            display: inline-block;
            background: #007bff;
            color: white;
            padding: 5px 12px;
            border-radius: 20px;
            font-size: 0.8rem;
            margin: 5px 0;
        }}
        @media (max-width: 768px) {{
            .container {{ padding: 15px; }}
            .header h1 {{ font-size: 1.8rem; }}
            .stats-bar {{ flex-direction: column; gap: 10px; }}
        }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h1>📊 Data Export Report</h1>
            <p>📁 File: <strong>{filename}</strong></p>
            <p>🕒 Generated: {__import__('datetime').datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
        </div>
        
        <div class="stats-bar">
            <div class="stat-item">
                <div class="stat-number">{len(data)-1 if data else 0}</div>
                <div class="stat-label">Total Rows</div>
            </div>
            <div class="stat-item">
                <div class="stat-number">{len(data[0]) if data else 0}</div>
                <div class="stat-label">Columns</div>
            </div>
            <div class="stat-item">
                <div class="stat-number">{len(data[0]) * (len(data)-1) if data else 0}</div>
                <div class="stat-label">Data Points</div>
            </div>
        </div>
        
        {f'<div class="filter-badge">🔍 Filtered by: "{search_term}"</div>' if search_term else ''}
    
    <div class="table-container">
        <table class="table table-striped">
'''
        
        if data and len(data) > 0:
            # Add header
            html_content += '<thead><tr>'
            for cell in data[0]:
                html_content += f'<th>{str(cell)}</th>'
            html_content += '</tr></thead>\n<tbody>'
            
            # Add data rows
            for row in data[1:]:
                html_content += '<tr>'
                for cell in row:
                    html_content += f'<td>{str(cell)}</td>'
                html_content += '</tr>'
            
            html_content += '</tbody>'
        
        html_content += '''
        </table>
    </div>
    
        <div class="summary">
            <h5>📋 Export Summary</h5>
            <p><strong>📊 Dataset:</strong> {len(data)-1 if data else 0} records across {len(data[0]) if data else 0} fields</p>
            <p><strong>💾 File Size:</strong> Approximately {round((len(str(data)) / 1024), 2)} KB</p>
            <p><strong>🎯 Usage:</strong> This table is fully scrollable horizontally and vertically. Use Ctrl+F to search within the page.</p>
            <p><strong>📱 Responsive:</strong> Optimized for desktop and mobile viewing</p>
        </div>
    </div>
    
    <script>
        // Add search functionality
        document.addEventListener('DOMContentLoaded', function() {{
            console.log('Data loaded successfully with {len(data)-1 if data else 0} rows');
        }});
    </script>
</body>
</html>
'''
        
        # Return HTML file
        buffer = BytesIO(html_content.encode('utf-8'))
        html_filename = f"{filename.split('.')[0]}_data.html"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=html_filename,
            mimetype='text/html'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error generating HTML: {str(e)}'})

@app.route('/download_doc', methods=['POST'])
def download_doc():
    filename = request.json.get('filename')
    search_term = request.json.get('search', '').lower()
    
    if not filename:
        return jsonify({'error': 'No filename provided'})
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    try:
        if filename.endswith('.xlsx'):
            data = read_excel_simple(filepath)
        else:
            data = read_csv_simple(filepath)
        
        # Apply search filter
        if search_term and data:
            filtered_data = [data[0]]  # Keep header
            for row in data[1:]:
                if any(search_term in str(cell).lower() for cell in row):
                    filtered_data.append(row)
            data = filtered_data
        
        # Create DOC file (HTML format that opens in Word)
        doc_content = f'''
<html>
<head>
    <meta charset="UTF-8">
    <title>📊 Data Export - {filename}</title>
    <style>
        body {{ font-family: Arial, sans-serif; margin: 20px; line-height: 1.6; }}
        .header {{ text-align: center; margin-bottom: 30px; padding: 20px; background: #f0f8ff; border-radius: 10px; }}
        h1 {{ color: #2c3e50; margin: 0; font-size: 28px; }}
        .info {{ background: #ecf0f1; padding: 15px; border-radius: 8px; margin: 20px 0; border-left: 4px solid #3498db; }}
        table {{ width: 100%; border-collapse: collapse; margin: 20px 0; font-size: 12px; }}
        th {{ background-color: #3498db; color: white; padding: 12px 8px; text-align: left; font-weight: bold; }}
        td {{ padding: 8px; border: 1px solid #bdc3c7; }}
        tr:nth-child(even) {{ background-color: #f8f9fa; }}
        .summary {{ background: #d5f4e6; padding: 15px; border-radius: 8px; margin-top: 20px; border-left: 4px solid #28a745; }}
        .summary h3 {{ color: #155724; margin-top: 0; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>📊 Data Export Report</h1>
        <p style="font-size: 16px; color: #666; margin: 10px 0;">Professional Data Analysis Document</p>
    </div>
    
    <div class="info">
        <p><strong>📁 Source File:</strong> {filename}</p>
        <p><strong>📅 Export Date:</strong> {__import__('datetime').datetime.now().strftime('%B %d, %Y at %I:%M %p')}</p>
        <p><strong>📊 Records:</strong> {len(data)-1 if data else 0} rows, {len(data[0]) if data else 0} columns</p>
        <p><strong>💾 Data Points:</strong> {len(data[0]) * (len(data)-1) if data else 0} total entries</p>
        {f'<p><strong>🔍 Filter Applied:</strong> "{search_term}"</p>' if search_term else ''}
    </div>
    
    <table>
'''
        
        if data and len(data) > 0:
            # Add header
            doc_content += '<thead><tr>'
            for cell in data[0]:
                doc_content += f'<th>{str(cell)}</th>'
            doc_content += '</tr></thead>\n<tbody>'
            
            # Add data rows
            for row in data[1:]:
                doc_content += '<tr>'
                for cell in row:
                    doc_content += f'<td>{str(cell)}</td>'
                doc_content += '</tr>'
            
            doc_content += '</tbody>'
        
        doc_content += f'''
    </table>
    
    <div class="summary">
        <h3>📋 Export Summary</h3>
        <p>✅ Successfully exported {len(data)-1 if data else 0} records</p>
        <p>📊 Total data points: {len(data[0]) * (len(data)-1) if data else 0}</p>
        <p>💼 This document can be opened in Microsoft Word, Google Docs, or any word processor</p>
        <p>🔄 Compatible with all major document editing software</p>
    </div>
</body>
</html>
'''
        
        # Return DOC file (HTML format)
        buffer = BytesIO(doc_content.encode('utf-8'))
        doc_filename = f"{filename.split('.')[0]}_data.doc"
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=doc_filename,
            mimetype='application/msword'
        )
        
    except Exception as e:
        return jsonify({'error': f'Error generating DOC: {str(e)}'})

if __name__ == '__main__':
    app.run(debug=True, port=5001)